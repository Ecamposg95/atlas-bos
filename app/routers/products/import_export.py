"""``GET /api/products/export/excel`` and ``POST /api/products/upload``
— bulk Excel/CSV catalog import-export.

Sprint 5b split — extracted verbatim from the original ``products.py``.
"""
from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from typing import Optional
from decimal import Decimal
from datetime import datetime
import io
import csv

from app.core.database import get_db
from app.core.tenant_context import get_current_active_organization
from app.models import (
    Product, ProductVariant, StockOnHand, User,
    InventoryMovement, MovementType, Department, ProductPrice,
    PackagingUnit, Brand, ProductBranchStatus,
)
from app.models.organization import Branch, BranchType
from app.core.security import get_current_user
from app.crud.products import query_visible_products, _is_admin

from ._shared import _safe_str, _is_na, _safe_decimal

router = APIRouter()


# -----------------------------
# 6. Exportar Excel (Template con productos existentes)
# -----------------------------
@router.get("/export/excel")
def export_products_template(
    branch_id: Optional[int] = None,
    search: Optional[str] = None,
    department_id: Optional[str] = None,
    brand_id: Optional[str] = None,
    approval_status: Optional[str] = None,
    include_inactive: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    org_id: int = Depends(get_current_active_organization)
):
    """
    Generates an Excel template with existing products for bulk update/import.

    Acepta los mismos filtros que `GET /products/` para que el "Exportar"
    del Admin Catalog descargue exactamente el subconjunto que el admin
    está viendo en pantalla (no el catálogo completo).

    [A1-08] Usa query_visible_products como fuente — CAJERO/GERENTE no
    pueden exportar catálogo cross-branch vía este endpoint.
    """
    import io
    from fastapi.responses import StreamingResponse

    # Track 3: cajero puede exportar SU sucursal con include_inactive.
    # branch_override (drill-down a otra sucursal) sigue admin-only —
    # query_visible_products limita a la sucursal del usuario para no-admin.
    branch_override = None
    if _is_admin(current_user) and branch_id and branch_id != 0:
        branch_override = branch_id

    q = query_visible_products(
        db,
        current_user,
        org_id,
        include_inactive=include_inactive,
        search=search or None,
        branch_id_override=branch_override,
    ).options(
        joinedload(Product.variants).joinedload(ProductVariant.prices),
        joinedload(Product.variants).joinedload(ProductVariant.packaging_units),
        joinedload(Product.department),
    ).filter(Product.deleted_at == None)

    if department_id:
        q = q.filter(Product.department_id == department_id)
    if brand_id:
        q = q.filter(Product.brand_id == brand_id)
    if approval_status and approval_status != 'ALL':
        q = q.filter(Product.approval_status == approval_status)

    query = q.all()

    data = []

    # 1. Resolve Target Branch (Explicit OR User's Branch OR Org's HQ)
    target_branch_id = branch_id
    if not target_branch_id:
        target_branch_id = current_user.branch_id

    if not target_branch_id:
        hq = db.query(Branch).filter(
            Branch.organization_id == org_id,
            Branch.branch_type == BranchType.HQ
        ).first()
        if hq:
            target_branch_id = hq.id
        else:
            # Fallback to any branch if HQ not found
            any_b = db.query(Branch).filter(Branch.organization_id == org_id).first()
            if any_b: target_branch_id = any_b.id

    # 2. Add existing products
    for p in query:
        if not p.variants: continue
        v = p.variants[0]

        # Get Stock
        stock_qty = 0
        if target_branch_id:
             s = db.query(StockOnHand).filter(
                 StockOnHand.variant_id == v.id,
                 StockOnHand.branch_id == target_branch_id
             ).first()
             if s: stock_qty = s.qty_on_hand

        row = {
            "SKU": v.sku,
            "Nombre": p.name,
            "Descripcion": p.description or "",
            "Unidad": p.unit,
            "Departamento": p.department.name if p.department else "General",
            "Marca": p.brand.name if p.brand else "",
            "Codigo Barras": v.barcode or "",
            "Precio Base": float(v.price),
            "Costo": float(v.cost),
            "Stock": float(stock_qty),
            "Incluye IVA": "Si" if v.has_iva else "No"
        }

        # Prices (up to 5)
        prices_sorted = sorted(v.prices, key=lambda x: x.min_quantity)
        for i, price in enumerate(prices_sorted[:5]):
            idx = i + 1
            row[f"P{idx} Nombre"] = price.price_name
            row[f"P{idx} Min"] = float(price.min_quantity)
            row[f"P{idx} Precio"] = float(price.unit_price)
            # Find linked pkg name if any
            pkg_name = ""
            if price.linked_package_id:
                # Naive search in this variant's packages
                 pkg = next((pk for pk in v.packaging_units if pk.id == price.linked_package_id), None)
                 if pkg: pkg_name = pkg.name
            row[f"P{idx} Empaque"] = pkg_name

        # Packaging (up to 3)
        pkgs_sorted = sorted(v.packaging_units, key=lambda x: x.units_per_package)
        for i, pkg in enumerate(pkgs_sorted[:3]):
            idx = i + 1
            row[f"E{idx} Nombre"] = pkg.name
            row[f"E{idx} Barcode"] = pkg.barcode or ""
            row[f"E{idx} Cantidad"] = float(pkg.units_per_package)
            row[f"E{idx} Precio"] = float(pkg.package_price)

        data.append(row)

    # 2. Build final column order (no pandas needed)
    standard_cols = [
        "SKU", "Nombre", "Descripcion", "Unidad", "Departamento", "Marca", "Codigo Barras",
        "Precio Base", "Costo", "Stock", "Incluye IVA"
    ]
    final_cols = standard_cols + \
                 [f"P{i} {f}" for i in range(1, 6) for f in ["Nombre", "Min", "Precio", "Empaque"]] + \
                 [f"E{i} {f}" for i in range(1, 4) for f in ["Nombre", "Barcode", "Cantidad", "Precio"]]

    # 3. Write Excel with openpyxl directly
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Plantilla'

    # Write header row
    worksheet.append(final_cols)

    # Write data rows (fill missing columns with empty string)
    for row_dict in data:
        worksheet.append([row_dict.get(col, "") for col in final_cols])

    # --- 1. DATA VALIDATION ---
    depts = [d.name for d in db.query(Department).filter(Department.organization_id == org_id).all()]
    brands = [b.name for b in db.query(Brand).filter(Brand.organization_id == org_id).all()]
    units = ["pza", "kg", "g", "ml", "L", "m", "cm", "caja", "pqte", "srv"]

    if depts or brands:
        val_sheet = workbook.create_sheet("Listas_Validacion")
        val_sheet.sheet_state = 'hidden'

        for i, d in enumerate(depts):
            val_sheet.cell(row=i+1, column=1, value=d)
        for i, b in enumerate(brands):
            val_sheet.cell(row=i+1, column=2, value=b)
        for i, u in enumerate(units):
            val_sheet.cell(row=i+1, column=3, value=u)

        max_row = max(1000, len(data) + 100)

        if depts:
            dv_dept = DataValidation(type="list", formula1=f"'Listas_Validacion'!$A$1:$A${len(depts)}", allow_blank=True)
            worksheet.add_data_validation(dv_dept)
            dv_dept.add(f"E2:E{max_row}")

        if brands:
            dv_brand = DataValidation(type="list", formula1=f"'Listas_Validacion'!$B$1:$B${len(brands)}", allow_blank=True)
            worksheet.add_data_validation(dv_brand)
            dv_brand.add(f"F2:F{max_row}")

        dv_unit = DataValidation(type="list", formula1=f"'Listas_Validacion'!$C$1:$C${len(units)}", allow_blank=True)
        worksheet.add_data_validation(dv_unit)
        dv_unit.add(f"D2:D{max_row}")

    # --- 2. STYLING ---
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        worksheet.column_dimensions[column].width = max(10, min(max_length + 2, 50))

    workbook.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    headers = {
        'Content-Disposition': f'attachment; filename="catalogo_productos_{timestamp}.xlsx"'
    }
    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

@router.post("/upload")
async def upload_products(
    branch_id: Optional[int] = Form(None),
    target_branch_ids: Optional[str] = Form(None), # Comma separated list
    scope: str = Form("branch"),  # "branch" | "all" | "selected"
    dry_run: bool = Form(False),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    org_id: int = Depends(get_current_active_organization),
):
    filename = (file.filename or "").lower()
    is_csv = filename.endswith(".csv")
    is_excel = filename.endswith((".xlsx", ".xlsm", ".xls"))

    if not (is_csv or is_excel):
        raise HTTPException(status_code=400, detail=f"Formato inválido. Use Excel o CSV.")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Archivo vacío.")

    try:
        if is_csv:
            text = contents.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            raw_rows = list(reader)
            rows = [{k.lower().strip(): v for k, v in r.items()} for r in raw_rows]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").lower().strip() for cell in ws[1]]
            rows = []
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: val for i, val in enumerate(row_cells)})
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo no se pudo leer. Verifique que sea un Excel o CSV válido e intente de nuevo.")

    # Cache departments and brands only for this org
    dept_map = {(c.name or "").lower(): c.id for c in db.query(Department).filter(Department.organization_id == org_id).all() if c.name}
    brand_map = {(b.name or "").lower(): b.id for b in db.query(Brand).filter(Brand.organization_id == org_id).all() if b.name}

    col_keys = set(rows[0].keys()) if rows else set()
    has_pkg_cols = any(k.startswith("e1") for k in col_keys)
    has_tier_cols = any(k.startswith("p1") for k in col_keys)

    created_count = 0
    updated_count = 0
    failed_count = 0
    failed_details = []
    preview_rows: list[dict] = []

    # Resolve Target Branch for this import (Explicit OR User's Branch OR Org's HQ)
    # RBAC: Si el usuario tiene sucursal asignada (CAJERO/GERENTE), forzar su sucursal
    # ignorando cualquier branch_id enviado en el form — impide afectar otras sucursales.
    if current_user.branch_id:
        target_branch_id = current_user.branch_id
    else:
        target_branch_id = branch_id

    # Validate explicit branch_id belongs to the user's organization (multi-tenancy guard)
    if target_branch_id and not current_user.branch_id:
        branch_check = db.query(Branch).filter(
            Branch.id == target_branch_id,
            Branch.organization_id == org_id
        ).first()
        if not branch_check:
            raise HTTPException(status_code=400, detail="Sucursal de destino inválida o no pertenece a tu organización.")

    if not target_branch_id:
        target_branch_id = current_user.branch_id

    if not target_branch_id:
        hq = db.query(Branch).filter(
            Branch.organization_id == org_id,
            Branch.branch_type == BranchType.HQ
        ).first()
        if hq:
            target_branch_id = hq.id
        else:
            # Fallback
            any_b = db.query(Branch).filter(Branch.organization_id == org_id).first()
            if any_b: target_branch_id = any_b.id

    for row in rows:
        try:
            with db.begin_nested():
                # Check basic fields
                raw_sku = _safe_str(row.get("sku")).strip()
                # Handle Excel auto-formatting issues (e.g. 1.0 -> "1")
                if raw_sku.endswith(".0"): raw_sku = raw_sku[:-2]
                raw_name = _safe_str(row.get("nombre"))

                if not raw_sku:
                    continue # Skip truly empty rows

                # Find or Create Product logic
                # Using SKU as key - SCOPED TO ORG
                existing_variant = (
                    db.query(ProductVariant)
                    .join(Product)
                    .filter(
                        ProductVariant.sku == raw_sku,
                        Product.organization_id == org_id,
                        ProductVariant.deleted_at == None
                    )
                    .first()
                )

                # --- PREPARE DATA ---
                # Department
                raw_dept = _safe_str(row.get("departamento", "General"))
                dept_id = dept_map.get(raw_dept.lower())
                if not dept_id and raw_dept:
                    new_dept = Department(name=raw_dept, organization_id=org_id)
                    db.add(new_dept)
                    db.flush()
                    dept_id = new_dept.id
                    dept_map[raw_dept.lower()] = dept_id

                # Brand
                raw_brand = _safe_str(row.get("marca", ""))
                brand_id = brand_map.get(raw_brand.lower())
                if not brand_id and raw_brand:
                    new_brand = Brand(name=raw_brand, organization_id=org_id)
                    db.add(new_brand)
                    db.flush()
                    brand_id = new_brand.id
                    brand_map[raw_brand.lower()] = brand_id

                price_base = _safe_decimal(row.get("precio base", row.get("precio", 0)))
                cost = _safe_decimal(row.get("costo", 0))
                raw_iva = _safe_str(row.get("incluye iva", "no")).strip().lower()
                has_iva_val = raw_iva in ("si", "sí", "yes", "1", "true")
                is_new = False

                if existing_variant:
                    # Update
                    prod = existing_variant.product
                    prod.name = raw_name if raw_name else prod.name
                    prod.department_id = dept_id
                    if brand_id:
                        prod.brand_id = brand_id
                    raw_desc = _safe_str(row.get("descripcion", ""))
                    if raw_desc:
                        prod.description = raw_desc
                    raw_unit = _safe_str(row.get("unidad", ""))
                    if raw_unit:
                        prod.unit = raw_unit
                    existing_variant.has_iva = has_iva_val
                    existing_variant.price = price_base
                    existing_variant.cost = cost
                    raw_barcode = _safe_str(row.get("codigo barras", ""))
                    if raw_barcode:
                        existing_variant.barcode = raw_barcode
                    updated_count += 1
                    variant = existing_variant
                    if len(preview_rows) < 20:
                        preview_rows.append({
                            "action": "UPDATE",
                            "sku": raw_sku,
                            "name": raw_name or prod.name,
                            "price": float(price_base) if price_base is not None else None,
                            "error_message": None,
                        })
                else:
                    # Create
                    if not raw_name: raw_name = "Producto sin Nombre"


                    # Determine Status based on User Role
                    # Always APPROVED as per user request (so cashiers can see what they import immediately)
                    initial_status = "APPROVED"

                    prod = Product(
                        name=raw_name,
                        description=_safe_str(row.get("descripcion", "")),
                        unit=_safe_str(row.get("unidad", "pza")),
                        department_id=dept_id,
                        brand_id=brand_id,
                        has_variants=True,
                        is_active=True,
                        organization_id=org_id,
                        approval_status=initial_status,
                        created_by_user_id=current_user.id
                    )
                    db.add(prod)
                    db.flush()

                    variant = ProductVariant(
                        product_id=prod.id,
                        sku=raw_sku,
                        barcode=_safe_str(row.get("codigo barras")),
                        variant_name="Estándar",
                        price=price_base,
                        cost=cost,
                        has_iva=has_iva_val,
                        organization_id=org_id
                    )
                    db.add(variant)
                    db.flush()
                    created_count += 1
                    is_new = True
                    if len(preview_rows) < 20:
                        preview_rows.append({
                            "action": "NEW",
                            "sku": raw_sku,
                            "name": raw_name,
                            "price": float(price_base) if price_base is not None else None,
                            "error_message": None,
                        })

                    # [MÓDULO 2] La habilitación comercial se maneja unificada abajo

                # [MÓDULO 2] Asegurar Habilitación Comercial (Context Aware)
                effective_branch_ids: set = set()

                if current_user.branch_id:
                    # Usuarios de sucursal (CAJERO/GERENTE): siempre su propia sucursal
                    effective_branch_ids.add(current_user.branch_id)
                else:
                    # Usuario HQ: usar scope para determinar alcance
                    if scope == "all":
                        # Todas las sucursales activas de la organización
                        org_branches = db.query(Branch).filter(
                            Branch.organization_id == org_id
                        ).all()
                        effective_branch_ids.update(b.id for b in org_branches)
                    elif scope == "selected" and target_branch_ids:
                        try:
                            ids = [int(x.strip()) for x in target_branch_ids.split(",") if x.strip()]
                            effective_branch_ids.update(ids)
                        except Exception:
                            failed_count += 1
                            failed_details.append(f"target_branch_ids inválido: '{target_branch_ids}'")
                            if len(preview_rows) < 20:
                                preview_rows.append({
                                    "action": "ERROR",
                                    "sku": (row.get("sku") or row.get("código") or row.get("codigo") or None),
                                    "name": row.get("nombre") or row.get("name"),
                                    "price": None,
                                    "error_message": f"target_branch_ids inválido: '{target_branch_ids}'",
                                })
                            continue
                    else:
                        # scope == "branch": usar HQ branch como fallback
                        hq_b = db.query(Branch).filter(
                            Branch.organization_id == org_id,
                            Branch.branch_type == BranchType.HQ
                        ).first()
                        if hq_b:
                            effective_branch_ids.add(hq_b.id)

                # Validar que todas las branch_ids seleccionadas pertenecen a la org
                if effective_branch_ids:
                    valid_ids = {
                        b.id for b in db.query(Branch).filter(
                            Branch.organization_id == org_id,
                            Branch.id.in_(effective_branch_ids)
                        ).all()
                    }
                    invalid = effective_branch_ids - valid_ids
                    if invalid:
                        failed_count += 1
                        failed_details.append(f"Sucursales no válidas para esta organización: {invalid}")
                        if len(preview_rows) < 20:
                            preview_rows.append({
                                "action": "ERROR",
                                "sku": (row.get("sku") or row.get("código") or row.get("codigo") or None),
                                "name": row.get("nombre") or row.get("name"),
                                "price": None,
                                "error_message": f"Sucursales no válidas para esta organización: {invalid}",
                            })
                        continue

                for bid in effective_branch_ids:
                    existing_status = db.query(ProductBranchStatus).filter(
                        ProductBranchStatus.variant_id == variant.id,
                        ProductBranchStatus.branch_id == bid
                    ).first()

                    if not existing_status:
                        db.add(ProductBranchStatus(
                            variant_id=variant.id,
                            branch_id=bid,
                            is_active_pos=True,
                            is_active_hq=True if not current_user.branch_id else False,
                            is_visible=True,
                            organization_id=org_id
                        ))
                    else:
                        existing_status.is_visible = True

                # --- STOCK ---
                input_stock = _safe_decimal(row.get("stock", -1)) # -1 sentinel

                if input_stock >= 0 and target_branch_id:
                    # Check current stock
                    stock_record = db.query(StockOnHand).filter(
                        StockOnHand.variant_id == variant.id,
                        StockOnHand.branch_id == target_branch_id
                    ).first()
                    if not stock_record:
                        stock_record = StockOnHand(
                            branch_id=target_branch_id,
                            variant_id=variant.id,
                            qty_on_hand=0,
                            is_active=True,
                            organization_id=org_id
                        )
                        db.add(stock_record)
                        db.flush()

                        # [REMOVED DUPLICATE BRANCH STATUS CHECK HERE] -> Done above

                    diff = input_stock - stock_record.qty_on_hand
                    if diff != 0:
                        stock_record.qty_on_hand = input_stock
                        reason = "Carga Masiva (Inicial)" if is_new else "Carga Masiva (Ajuste)"
                        db.add(InventoryMovement(
                            branch_id=target_branch_id,
                            variant_id=variant.id,
                            user_id=current_user.id,
                            movement_type=MovementType.ADJUSTMENT_IN if diff > 0 else MovementType.ADJUSTMENT_OUT,
                            qty_change=abs(diff),
                            qty_before=stock_record.qty_on_hand - diff,
                            qty_after=input_stock,
                            reference="Excel Upload",
                            notes=reason,
                            organization_id=org_id
                        ))

                # --- SAFE CLEANUP (DEPENDENCY HANDLING) ---
                if has_pkg_cols:
                    if has_tier_cols:
                        # Will replace prices too, so delete them first
                        db.query(ProductPrice).filter(ProductPrice.variant_id == variant.id).delete()
                    else:
                        # Not replacing prices, but deleting packages they might link to. Unlink them.
                        db.query(ProductPrice).filter(ProductPrice.variant_id == variant.id).update({ProductPrice.linked_package_id: None})

                    # Delete packages
                    db.query(PackagingUnit).filter(PackagingUnit.variant_id == variant.id).delete()
                elif has_tier_cols:
                    # Replacing prices only
                    db.query(ProductPrice).filter(ProductPrice.variant_id == variant.id).delete()


                # --- RECREATE PACKAGING UNITS ---
                added_pkgs = {}
                if has_pkg_cols:
                    for i in range(1, 4): # Up to 3
                        e_name = row.get(f"e{i} nombre")
                        e_barcode = row.get(f"e{i} barcode")
                        e_qty = row.get(f"e{i} cantidad")
                        e_price = row.get(f"e{i} precio")

                        if e_name and not _is_na(e_name) and e_price and not _is_na(e_price):
                            pkg = PackagingUnit(
                                variant_id=variant.id,
                                name=str(e_name),
                                barcode=str(e_barcode) if not _is_na(e_barcode) else None,
                                units_per_package=_safe_decimal(e_qty, 1),
                                package_price=_safe_decimal(e_price, 0),
                                organization_id=org_id
                            )
                            db.add(pkg)
                            added_pkgs[str(e_name).lower()] = pkg

                    db.flush() # To get IDs

                # --- RECREATE TIERED PRICES ---
                if has_tier_cols:
                    # Note: We already deleted them in the cleanup phase
                    for i in range(1, 6): # Up to 5
                        p_name = row.get(f"p{i} nombre")
                        p_min = row.get(f"p{i} min")
                        p_val = row.get(f"p{i} precio")
                        p_link = row.get(f"p{i} empaque")

                        if p_name and not _is_na(p_name) and p_val and not _is_na(p_val):
                            linked_pkg_id = None
                            target_link = str(p_link).lower() if p_link and not _is_na(p_link) else None

                            # Auto-link if name matches or explicit link
                            if target_link and target_link in added_pkgs:
                                linked_pkg_id = added_pkgs[target_link].id
                            elif str(p_name).lower() in added_pkgs:
                                linked_pkg_id = added_pkgs[str(p_name).lower()].id

                            db.add(ProductPrice(
                                variant_id=variant.id,
                                price_name=str(p_name),
                                min_quantity=_safe_decimal(p_min, 1),
                                unit_price=_safe_decimal(p_val, 0),
                                linked_package_id=linked_pkg_id,
                                organization_id=org_id
                            ))

        except Exception as e:
            err_msg = str(e)
            print(f"Error row {idx}: {err_msg}")
            failed_count += 1
            failed_details.append({"row": idx + 2, "sku": str(row.get("sku", "N/A")), "error": err_msg})
            if len(preview_rows) < 20:
                preview_rows.append({
                    "action": "ERROR",
                    "sku": (row.get("sku") or row.get("código") or row.get("codigo") or None),
                    "name": row.get("nombre") or row.get("name"),
                    "price": None,
                    "error_message": err_msg,
                })

        # Batch Commit to avoid holding lock too long
        # Skip batch commits in dry_run so we can rollback all changes at the end.
        if not dry_run and (created_count + updated_count + failed_count) % 50 == 0:
             try:
                db.commit()
             except Exception as commit_err:
                db.rollback()
                print(f"Batch commit error: {commit_err}")

    if dry_run:
        db.rollback()
        return {
            "total_rows": created_count + updated_count + failed_count,
            "to_create": created_count,
            "to_update": updated_count,
            "errors": failed_count,
            "preview": preview_rows,
            "error_details": failed_details,
        }

    # Final Commit for remaining items
    try:
        db.commit()
    except Exception as final_commit_err:
        db.rollback()
        print(f"Final commit error: {final_commit_err}")

    # Calculate pending based on created (if status was pending) for feedback
    # Since we auto-approve all imports now, pending is always 0
    pending_count = 0

    return {
        "created": created_count,
        "updated": updated_count,
        "failed": failed_count,
        "pending": pending_count,
        "details": failed_details
    }
